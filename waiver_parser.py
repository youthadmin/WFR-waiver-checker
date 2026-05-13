#!/usr/bin/env python3
"""
waiver_parser.py — extract a WaiverPerson list from a waiver report.

Primary input: PDF (the live WFR weekly email attaches a PDF). pdfplumber
handles table extraction; if no tables are detected the parser falls back
to line-by-line text extraction. If pdfplumber ever struggles with the
actual WFR PDF format, swap in pypdf or pdfminer.six in `parse_pdf` and
document the reason inline.

Secondary inputs (CSV, XLSX, HTML email body) are supported lightly — the
shapes are correct but they're untested against real samples since we
don't have any.

Name normalization handles:
  • "First Last" and "Last, First" orderings
  • Honorifics: Mr / Mrs / Ms / Miss / Dr / Rev / Fr (with or without dot)
  • Suffixes: Jr / Sr / II / III / IV
  • Middle names / initials (dropped from match name, preserved in raw_name)
  • Hyphens and apostrophes in surnames (Mary-Kate, O'Brien)

Run standalone for a smoke test:
    python waiver_parser.py fixtures/sample_waiver_simple.pdf
"""

from __future__ import annotations

import csv
import io
import re
import sys
from pathlib import Path
from typing import Optional, Union

from pco_matcher import WaiverPerson  # canonical dataclass — single source of truth

HONORIFICS = {"mr", "mrs", "ms", "miss", "dr", "rev", "fr"}
SUFFIXES = {"jr", "sr", "ii", "iii", "iv"}

_DATE_TAIL_RE = re.compile(r"\s+(\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4}).*$")


def _norm_token(tok: str) -> str:
    return tok.rstrip(",.").lower()


def normalize_name(raw: str) -> Optional[WaiverPerson]:
    """Parse a single name string into a WaiverPerson.

    Returns None if no usable name can be recovered.
    """
    raw = (raw or "").strip().strip('"\'')
    if not raw:
        return None

    # Strip trailing date columns that sometimes survive text extraction
    raw_for_split = _DATE_TAIL_RE.sub("", raw).strip()
    if not raw_for_split:
        return None

    if "," in raw_for_split:
        # "Last [Suffix], First [Middle…]"
        last_part, first_part = (p.strip() for p in raw_for_split.split(",", 1))
        first_tokens = [t for t in first_part.split() if _norm_token(t) not in HONORIFICS]
        if not first_tokens:
            return None
        first = first_tokens[0]
        last_tokens = [t for t in last_part.split() if _norm_token(t) not in SUFFIXES]
        last = " ".join(last_tokens) if last_tokens else last_part
    else:
        # "[Honorific] First [Middle…] Last [Suffix]"
        tokens = raw_for_split.split()
        while tokens and _norm_token(tokens[0]) in HONORIFICS:
            tokens.pop(0)
        while tokens and _norm_token(tokens[-1]) in SUFFIXES:
            tokens.pop()
        if not tokens:
            return None
        if len(tokens) == 1:
            first, last = tokens[0], ""
        else:
            first, last = tokens[0], tokens[-1]

    first = first.strip()
    last = last.strip()
    if not (first or last):
        return None
    return WaiverPerson(first_name=first, last_name=last, raw_name=raw, email=None)


def _looks_like_header_or_meta(line: str) -> bool:
    lower = line.lower().strip()
    if not lower:
        return True
    prefixes = ("camp:", "generated", "report", "wfr ", "page ", "total")
    if lower.startswith(prefixes):
        return True
    if "waiver" in lower and ("complet" in lower or "report" in lower):
        return True
    if lower in {"name", "camper name", "participant", "attendee", "date completed", "completed"}:
        return True
    # Concatenated header rows like "Camper Name Date Completed"
    if "name" in lower and ("completed" in lower or "date" in lower):
        return True
    return False


def _extract_names_from_table(table: list[list[Optional[str]]]) -> list[str]:
    """Pick the name column(s) from a 2-D table and return cell strings."""
    if not table:
        return []
    header = [(c or "").strip().lower() for c in table[0]]
    has_header = any(header)

    first_col = last_col = name_col = None
    for i, h in enumerate(header):
        if "first" in h and "name" in h:
            first_col = i
        elif "last" in h and "name" in h:
            last_col = i
        elif "name" in h or "camper" in h or "participant" in h or "attendee" in h:
            if name_col is None:
                name_col = i

    rows = table[1:] if has_header else table

    if first_col is not None and last_col is not None:
        return [
            f"{(r[first_col] or '').strip()} {(r[last_col] or '').strip()}".strip()
            for r in rows
            if r and (len(r) > max(first_col, last_col)) and (r[first_col] or r[last_col])
        ]
    if name_col is not None:
        return [(r[name_col] or "").strip() for r in rows if r and len(r) > name_col and r[name_col]]
    # No header — assume first non-empty column is the name
    return [(r[0] or "").strip() for r in rows if r and r[0]]


def _dedupe_append(results: list[WaiverPerson], seen: set, wp: Optional[WaiverPerson]) -> bool:
    if wp is None or not (wp.first_name or wp.last_name):
        return False
    key = (wp.first_name.lower(), wp.last_name.lower())
    if key in seen:
        return False
    seen.add(key)
    results.append(wp)
    return True


def parse_pdf(source: Union[str, Path, io.IOBase]) -> list[WaiverPerson]:
    """Tables first, then per-line text fallback. pdfplumber handles both."""
    import pdfplumber

    results: list[WaiverPerson] = []
    seen: set = set()

    with pdfplumber.open(source) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            if tables:
                # Trust table extraction even if it yields zero data rows
                # (e.g. a header-only table on an empty-week report).
                for table in tables:
                    for cell in _extract_names_from_table(table):
                        _dedupe_append(results, seen, normalize_name(cell))
                continue

            # No tables detected — line-by-line text fallback for plain rosters.
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line or _looks_like_header_or_meta(line):
                    continue
                wp = normalize_name(line)
                if wp is None or not wp.last_name:
                    # Single-token lines aren't useful matches; skip
                    continue
                _dedupe_append(results, seen, wp)

    return results


def _from_dict_rows(rows: list[dict]) -> list[WaiverPerson]:
    results: list[WaiverPerson] = []
    seen: set = set()
    for row in rows:
        row_lower = {(k or "").strip().lower(): (v or "") for k, v in row.items()}
        first = (row_lower.get("first name") or row_lower.get("first") or "").strip()
        last = (row_lower.get("last name") or row_lower.get("last") or "").strip()
        full = (row_lower.get("camper name") or row_lower.get("name")
                or row_lower.get("attendee") or row_lower.get("participant") or "").strip()
        email = (row_lower.get("email") or "").strip() or None
        if first or last:
            wp = normalize_name(f"{first} {last}".strip())
        elif full:
            wp = normalize_name(full)
        else:
            continue
        if wp is None:
            continue
        if email:
            wp.email = email.lower()
        _dedupe_append(results, seen, wp)
    return results


def parse_csv(source: Union[str, Path]) -> list[WaiverPerson]:
    path = Path(source)
    if path.exists():
        with open(path, newline="", encoding="utf-8-sig") as f:
            return _from_dict_rows(list(csv.DictReader(f)))
    # Treat as raw CSV text
    return _from_dict_rows(list(csv.DictReader(io.StringIO(str(source)))))


def parse_xlsx(source: Union[str, Path]) -> list[WaiverPerson]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=str(source), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter, [])]
    dict_rows = [
        dict(zip(header, [(str(c).strip() if c is not None else "") for c in row]))
        for row in rows_iter
    ]
    return _from_dict_rows(dict_rows)


def parse_html(html_text: str) -> list[WaiverPerson]:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html_text, "lxml")
    results: list[WaiverPerson] = []
    seen: set = set()

    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        # Try header detection from <th> or first row
        header_cells = rows[0].find_all(["th", "td"])
        header = [c.get_text(strip=True).lower() for c in header_cells]
        body = rows[1:] if any("name" in h for h in header) else rows
        dict_rows = []
        for r in body:
            cells = [c.get_text(strip=True) for c in r.find_all(["th", "td"])]
            dict_rows.append(dict(zip(header, cells)))
        for wp in _from_dict_rows(dict_rows):
            _dedupe_append(results, seen, wp)

    if results:
        return results

    # Fallback: line-by-line on stripped text
    text = soup.get_text("\n")
    for line in text.splitlines():
        line = line.strip()
        if not line or _looks_like_header_or_meta(line):
            continue
        wp = normalize_name(line)
        if wp is None or not wp.last_name:
            continue
        _dedupe_append(results, seen, wp)
    return results


def parse(source: Union[str, Path, bytes]) -> list[WaiverPerson]:
    """Auto-detect format and dispatch."""
    if isinstance(source, (str, Path)):
        path = Path(source)
        suffix = path.suffix.lower()
        if path.exists():
            if suffix == ".pdf":
                return parse_pdf(path)
            if suffix == ".csv":
                return parse_csv(path)
            if suffix in {".xlsx", ".xls"}:
                return parse_xlsx(path)
            if suffix in {".html", ".htm"}:
                return parse_html(path.read_text(encoding="utf-8"))
            with open(path, "rb") as f:
                head = f.read(8)
            if head.startswith(b"%PDF"):
                return parse_pdf(path)
            if head.startswith(b"PK\x03\x04"):
                return parse_xlsx(path)
            return parse_csv(path)
        # Not a path — treat as text payload (HTML or CSV)
        text = str(source)
        if "<" in text[:200]:
            return parse_html(text)
        return parse_csv(text)
    if isinstance(source, bytes):
        if source[:4] == b"%PDF":
            return parse_pdf(io.BytesIO(source))
        if source[:4] == b"PK\x03\x04":
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(source)
                return parse_xlsx(tmp.name)
        return parse_html(source.decode("utf-8", errors="replace"))
    raise TypeError(f"unsupported source type: {type(source)}")


def _smoke_test(path: str) -> int:
    print(f"Parsing {path}…")
    people = parse(path)
    print(f"Extracted {len(people)} name(s):")
    for p in people:
        email = f"  <{p.email}>" if p.email else ""
        print(f"  • first={p.first_name!r:18s}  last={p.last_name!r:20s}  raw={p.raw_name!r}{email}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python waiver_parser.py <file>", file=sys.stderr)
        sys.exit(2)
    sys.exit(_smoke_test(sys.argv[1]))
